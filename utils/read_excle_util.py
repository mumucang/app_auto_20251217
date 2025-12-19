import os

import openpyxl
from openpyxl.utils import get_column_letter

from config import datas_path


class ReaderExcelUtil:
    @staticmethod
    def read_excel(file_path,sheet_name):
        # 加载Excel文件
        workbook = openpyxl.load_workbook(file_path)
        # 选择第一个工作表
        sheet = workbook[sheet_name]

        # 获取所有行的数据
        data = []
        for row in sheet.iter_rows(min_row=2):
            # 将None值转换为空字符串
            row_data = tuple(cell.value if cell.value is not None else '' for cell in row)
            data.append(row_data)

        return data

# 使用示例
if __name__ == "__main__":
    file_path = os.path.join(datas_path, "app_datas.xlsx")  # 替换为你的Excel文件路径

    data = ReaderExcelUtil.read_excel(file_path,"Sheet1")
    print(data)
